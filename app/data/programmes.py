import pandas as pd
import sys
import os

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from app import create_app, db
from openpyxl import load_workbook

def process_programmes(file_path):
    file_name = os.path.basename(file_path)
    print(f"\nUpdating programmes from {file_name}")

    # Define the cell ranges and expected columns for each table
    table_ranges = {
        'PROGRAMMES': ('A1:E100', ['NAME', 'LEVEL1_CREDITS', 'ADVANCED_CREDITS', 'DEPARTMENT', 'NOTES']),
        'PROGRAMME_LEVEL1_COURSES': ('G1:H100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_ADVANCED_COURSES': ('J1:K100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_COURSE_ALTERNATIVES': ('M1:O100', ['PROGRAMME', 'COURSE', 'ALTERNATIVE']),
        'PROGRAMME_DEPARTMENTAL_REQUIREMENTS': ('Q1:T100', ['PROGRAMME', 'DEPARTMENT', 'CREDITS_REQUIRED', 'LEVEL']),
        'PROGRAMME_RECOMMENDED_COURSES': ('V1:W100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_OPTIONAL_ADVANCED_COURSES': ('Y1:Z100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_REQUIREMENT_FROM_OPTIONAL_ADVANCED_COURSES': ('AB1:AD100', ['PROGRAMME', 'NUMBER_OF_COURSES_REQUIRED', 'CREDITS_REQUIRED']),
        'PROGRAMME_OPTIONAL_CORE_COURSES_SET1': ('AF1:AG100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_OPTIONAL_CORE_COURSES_SET2': ('AI1:AJ100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_OPTIONAL_CORE_COURSES_SET3': ('AL1:AM100', ['PROGRAMME', 'COURSE']),
        'PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET1': ('AO1:AQ100', ['PROGRAMME', 'NUMBER_OF_COURSES_REQUIRED', 'CREDITS_REQUIRED']),
        'PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET2': ('AS1:AU100', ['PROGRAMME', 'NUMBER_OF_COURSES_REQUIRED', 'CREDITS_REQUIRED']),
        'PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET3': ('AW1:AY100', ['PROGRAMME', 'NUMBER_OF_COURSES_REQUIRED', 'CREDITS_REQUIRED']),
        'PROGRAMME_MANDATORY_SUMMER_COURSES': ('BA1:BB100', ['PROGRAMME', 'COURSE'])
    }

    def column_to_number(column_label):
        number = 0
        for i in range(len(column_label)):
            number = number * 26 + (ord(column_label[i]) - ord('A') + 1)
        return number

    # Load the Excel file
    wb = load_workbook(file_path)

    # Create an empty dictionary to store table data
    all_tables = {}

    # Get the 'Programmes' sheet
    try:
        sheet = wb['Programmes']
    except KeyError:
        print("The 'Programmes' sheet was not found in the Excel file.")
        return

    for table_name, (cell_range, expected_columns) in table_ranges.items():
        table_data = []
        start_row = int(''.join(filter(str.isdigit, cell_range.split(':')[0]))) + 2  # Extract the start row from the cell range and skip the first two rows
        start_column = column_to_number(cell_range.split(':')[0].upper().strip('0123456789'))  # Extract the start column from the cell range
        end_column = column_to_number(cell_range.split(':')[1].upper().strip('0123456789'))  # Extract the end column from the cell range

        # Initialize row counter
        row_counter = start_row

        # Check if all rows after the column headers are empty
        empty_table = all(all(sheet.cell(row=row_index, column=col).value is None for col in range(start_column, end_column + 1)) for row_index in range(start_row, sheet.max_row + 1))
        if empty_table:
            #print(f"Table {table_name} is empty.")
            continue  # Skip to the next table


        while True:
            row = [sheet.cell(row=row_counter, column=col).value for col in range(start_column, end_column + 1)]
            # Strip whitespace from cell values and check if they are None
            if all(cell is None or str(cell).strip() == '' for cell in row):
                break  # Stop processing when an empty row is encountered
            table_data.append(row)
            row_counter += 1

        # Store table data in the dictionary
        all_tables[table_name] = table_data

    # Process and insert data into the database
    app = create_app()

    # Process and insert data into the database
    app = create_app()

    with app.app_context():
        conn = db.engine.raw_connection()

        # Create a cursor object
        cur = conn.cursor()

        # Iterate through tables and insert into database
        for table_name, table_data in all_tables.items():
            #print(f"Processing table: {table_name}")

            inserted_rows = 0
            updated_rows = 0

            # Check if the table is in the table ranges
            if table_name in table_ranges:
                data_length = len(table_data)
                if data_length < 1:
                    #print(f"Warning: Table '{table_name}' has no data.")
                    continue

                # Insert data into the corresponding table in the database
                try:
                    for row in table_data:
                        result = insert_into_table(table_name, table_ranges[table_name][1], row, cur)
                        if result == 'inserted':
                            inserted_rows += 1
                        elif result == 'updated':
                            updated_rows += 1
                except Exception as e:
                    print(f"Error inserting data into table '{table_name}': {e}")

                print(f"Inserted {inserted_rows} rows and updated {updated_rows} rows in table '{table_name}'.")

        # Commit changes and close connection
        conn.commit()
        conn.close()

def insert_into_table(table_name, expected_columns, row, cur):
    primary_keys = {
        "PROGRAMMES": "NAME",
        "PROGRAMME_LEVEL1_COURSES": "PROGRAMME, COURSE",
        "PROGRAMME_ADVANCED_COURSES": "PROGRAMME, COURSE",
        "PROGRAMME_COURSE_ALTERNATIVES": "PROGRAMME, COURSE, ALTERNATIVE",
        "PROGRAMME_DEPARTMENTAL_REQUIREMENTS": "PROGRAMME, DEPARTMENT",
        "PROGRAMME_RECOMMENDED_COURSES": "PROGRAMME, COURSE",
        "PROGRAMME_OPTIONAL_ADVANCED_COURSES": "PROGRAMME, COURSE",
        "PROGRAMME_REQUIREMENT_FROM_OPTIONAL_ADVANCED_COURSES": "PROGRAMME",
        "PROGRAMME_OPTIONAL_CORE_COURSES_SET1": "PROGRAMME, COURSE",
        "PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET1": "PROGRAMME",
        "PROGRAMME_OPTIONAL_CORE_COURSES_SET2": "PROGRAMME, COURSE",
        "PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET2": "PROGRAMME",
        "PROGRAMME_OPTIONAL_CORE_COURSES_SET3": "PROGRAMME, COURSE",
        "PROGRAMME_REQUIREMENT_FROM_OPTIONAL_CORE_COURSES_SET3": "PROGRAMME",
    }

    # Construct the INSERT query
    columns_str = ', '.join(expected_columns)
    placeholders_str = ', '.join(['%s'] * len(row))
    query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders_str})"

    # Get the primary key for the table
    primary_key = primary_keys[table_name]

    # Construct the ON CONFLICT DO NOTHING clause
    query += f" ON CONFLICT ({primary_key}) DO NOTHING"

    # Execute the INSERT query with the row data
    cur.execute(query, row)

    # Check if a row was inserted
    if cur.rowcount == 1:
        return 'inserted'
    else:
        # If no row was inserted, it means a conflict occurred. Update the existing row.
        update_str = ', '.join([f"{column} = %s" for column in expected_columns])
        where_str = ' AND '.join([f"{column} = %s" for column in primary_key.split(', ')])
        query = f"UPDATE {table_name} SET {update_str} WHERE {where_str}"
        cur.execute(query, row + row)
        return 'updated'